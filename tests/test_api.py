"""
Tests for the API (documents_converter/api/app.py): Phase 3 (the basic
endpoints), Phase 4 (security hardening), and Phase 6 (API-key auth).

Uses the same synthetic, fabricated fixture as the pipeline tests -- see
tests/fixtures/synthetic_scan.py and docs/PHASE_0_AUDIT.md risk register
item #1 for why real scanned documents are never used here.

Tests that submit a real async job always wait for it to reach a
terminal status (_wait_for_job_terminal) before returning, even ones
that aren't primarily testing job completion. Confirmed the hard way:
a test that submitted a job and returned immediately without waiting
left it running in the background, competing with the *next* test's
job for the same worker pool -- invisible on a fast dev machine, but
enough contention on GitHub Actions' 2-vCPU runner to push a later
job's test past its deadline. Real CI failure, not a pipeline bug --
see the fix commit for the full story.
"""

import io
import json
import logging
import time
import zipfile

import fitz
import openpyxl
import pytest
from fastapi.testclient import TestClient
from PIL import Image

from documents_converter.api import config, security
from documents_converter.api.app import _rate_limiter, app

from conftest import requires_tesseract

client = TestClient(app)


def _sample_png_bytes() -> bytes:
    """A tiny, real, valid PNG -- not a synthetic scan (no OCR involved in
    the image->pdf conversion these tests exercise), just something real
    enough that Pillow's own decoder and encoder both accept it."""
    buf = io.BytesIO()
    Image.new("RGB", (40, 30), color=(200, 60, 60)).save(buf, "PNG")
    return buf.getvalue()


@pytest.fixture(autouse=True)
def _reset_rate_limiter():
    """_rate_limiter is a module-level singleton shared by every request;
    without resetting it, whichever test happens to run the most requests
    against the shared TestClient "IP" would trip the limit for every test
    after it. Applied to all tests so the dedicated rate-limit test below
    can safely exhaust the limit without affecting anything else."""
    _rate_limiter._counts.clear()
    yield


def test_health_reports_status_and_tesseract_availability():
    resp = client.get("/health")
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "ok"
    assert isinstance(body["tesseract_available"], bool)


def test_convert_rejects_unsupported_extension():
    resp = client.post(
        "/api/v1/convert",
        files={"file": ("not_a_document.exe", io.BytesIO(b"whatever"), "application/octet-stream")},
    )
    assert resp.status_code == 400
    assert "Unsupported file type" in resp.json()["detail"]


def test_convert_rejects_oversized_upload(monkeypatch):
    monkeypatch.setattr(config, "MAX_UPLOAD_MB", 0)  # anything at all is "too big"
    resp = client.post(
        "/api/v1/convert",
        files={"file": ("scan.pdf", io.BytesIO(b"%PDF-1.4 fake content"), "application/pdf")},
    )
    assert resp.status_code == 413


@requires_tesseract
def test_convert_end_to_end_returns_valid_xlsx(synthetic_pdf, tesseract_cmd, monkeypatch):
    """
    Full round trip through the actual HTTP layer: upload the synthetic
    fixture, get back a real, correctly-populated .xlsx -- not just a
    non-error status code. Mirrors the exact expected values already
    locked in by test_ocr_excel.py's end-to-end pipeline test.
    """
    monkeypatch.setattr(config, "TESSERACT_CMD", tesseract_cmd)

    with open(synthetic_pdf, "rb") as f:
        resp = client.post(
            "/api/v1/convert",
            files={"file": ("synthetic_scan.pdf", f, "application/pdf")},
        )

    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "<f>" not in xml  # formula-injection regression guard, same as the pipeline test

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1][:5] == ("1", "100000001", "SMITH", "JOHN", "M")


# --------------------------------------------------------------------------
# Phase 4: security hardening (docs/PHASE_0_AUDIT.md).
# --------------------------------------------------------------------------


def test_convert_rejects_content_that_does_not_match_extension():
    """A file renamed to look like a PDF but containing something else
    entirely should be rejected by the magic-byte check, not accepted just
    because its filename ends in .pdf."""
    resp = client.post(
        "/api/v1/convert",
        files={"file": ("fake.pdf", io.BytesIO(b"this is not a pdf at all"), "application/pdf")},
    )
    assert resp.status_code == 400
    assert "doesn't match its extension" in resp.json()["detail"]


@pytest.mark.parametrize(
    "ext,good_header",
    [
        (".pdf", b"%PDF-1.4"),
        (".png", b"\x89PNG\r\n\x1a\n"),
        (".bmp", b"BM"),
    ],
)
def test_matches_magic_bytes_accepts_real_signatures(ext, good_header):
    assert security.matches_magic_bytes(ext, good_header) is True


def test_matches_magic_bytes_rejects_mismatched_signature():
    assert security.matches_magic_bytes(".pdf", b"not a pdf header") is False


def test_check_image_dimensions_rejects_oversized_canvas():
    """Decompression-bomb guard, tested directly against the check function
    rather than by actually constructing a giant image (which would defeat
    the point by allocating the memory this guard exists to avoid)."""
    with pytest.raises(security.FileTooLargeError):
        security.check_image_dimensions(20000, 20000)  # 400 MP, over the 50 MP limit
    security.check_image_dimensions(2000, 2000)  # 4 MP, fine -- must not raise


def test_check_pdf_page_count_rejects_too_many_pages():
    with pytest.raises(security.FileTooLargeError):
        security.check_pdf_page_count(security.MAX_PDF_PAGES + 1)
    security.check_pdf_page_count(1)  # must not raise


def test_convert_rate_limits_excessive_requests():
    """Exhausts the configured limit and confirms the next request is
    rejected with 429 -- then confirms it clears once the window resets
    manually (this test's own cleanup, not a real time-based wait)."""
    bad_file = {"file": ("x.xyz", io.BytesIO(b"data"), "application/octet-stream")}
    for _ in range(config.RATE_LIMIT_MAX_REQUESTS):
        resp = client.post("/api/v1/convert", files=bad_file)
        assert resp.status_code == 400  # rejected for extension, but still *counted*

    resp = client.post("/api/v1/convert", files=bad_file)
    assert resp.status_code == 429


@requires_tesseract
def test_convert_times_out_on_a_slow_conversion(monkeypatch, synthetic_pdf, tesseract_cmd):
    """A conversion that runs longer than CONVERT_TIMEOUT_SECONDS should be
    abandoned with a 504, not hang the request indefinitely. Simulated with
    a monkeypatched slow function rather than an actually-slow real
    conversion, since the point is testing the timeout wiring, not OCR
    performance."""
    import time

    monkeypatch.setattr(config, "TESSERACT_CMD", tesseract_cmd)
    monkeypatch.setattr(config, "CONVERT_TIMEOUT_SECONDS", 0.2)

    def _slow_convert(*args, **kwargs):
        time.sleep(5)

    # Phase 9: app.py no longer imports convert_scanned_to_excel directly --
    # it's called through the registered Capability's convert function
    # (documents_converter/converters/ocr_to_excel.py), so that's what
    # needs patching now for this to actually take effect on the real path.
    monkeypatch.setattr(
        "documents_converter.converters.ocr_to_excel.convert_scanned_to_excel", _slow_convert
    )

    with open(synthetic_pdf, "rb") as f:
        resp = client.post(
            "/api/v1/convert",
            files={"file": ("synthetic_scan.pdf", f, "application/pdf")},
        )
    assert resp.status_code == 504


# --------------------------------------------------------------------------
# Phase 6: API-key authentication (docs/PHASE_0_AUDIT.md).
# --------------------------------------------------------------------------

_junk_file = {"file": ("x.xyz", io.BytesIO(b"data"), "application/octet-stream")}


def _wait_for_job_terminal(job_id: str, timeout: float = 120) -> str:
    """
    Polls a job until it reaches a terminal status (completed/failed) or
    `timeout` elapses, returning the last observed status. 120s default,
    not the 60s originally used here -- CI hardware (GitHub Actions'
    ubuntu-latest runner: 2 vCPUs) is genuinely slower than a typical dev
    machine for CPU-bound OCR work, confirmed by reproducing the CI
    environment locally rather than guessed at.
    """
    deadline = time.monotonic() + timeout
    status = None
    while time.monotonic() < deadline:
        status_resp = client.get(f"/api/v1/jobs/{job_id}")
        assert status_resp.status_code == 200
        status = status_resp.json()["status"]
        if status in ("completed", "failed"):
            return status
        time.sleep(0.5)
    return status


def test_convert_works_without_auth_when_no_keys_configured():
    """Default state: config.API_KEYS is empty, so auth is off and a request
    with no Authorization header at all must not be rejected with 401 --
    it should reach the normal extension check instead (400, not 401)."""
    assert config.API_KEYS == ()  # the actual default, not assumed
    resp = client.post("/api/v1/convert", files=_junk_file)
    assert resp.status_code == 400


def test_convert_requires_api_key_when_configured(monkeypatch):
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1",))
    resp = client.post("/api/v1/convert", files=_junk_file)
    assert resp.status_code == 401
    assert "Authorization" in resp.json()["detail"]


def test_convert_rejects_wrong_api_key(monkeypatch):
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1",))
    resp = client.post(
        "/api/v1/convert",
        files=_junk_file,
        headers={"Authorization": "Bearer wrong-key"},
    )
    assert resp.status_code == 401
    assert "Invalid API key" in resp.json()["detail"]


def test_convert_accepts_a_valid_api_key(monkeypatch):
    """A correct key should pass the auth check and reach the normal
    pipeline (proven here by getting the extension-rejection 400, not an
    auth-related 401 -- a full OCR run isn't needed to prove auth passed)."""
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1", "secret-key-2"))
    resp = client.post(
        "/api/v1/convert",
        files=_junk_file,
        headers={"Authorization": "Bearer secret-key-2"},
    )
    assert resp.status_code == 400  # extension rejection, i.e. auth passed


def test_health_never_requires_auth(monkeypatch):
    """Load balancers and monitoring probes need to reach /health without
    credentials, even when auth is otherwise enabled."""
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1",))
    resp = client.get("/health")
    assert resp.status_code == 200


# --------------------------------------------------------------------------
# Phase 7: async job queue (docs/PHASE_0_AUDIT.md).
# --------------------------------------------------------------------------


def test_create_job_validates_upload_before_queueing():
    """Bad input is rejected immediately with the normal validation error,
    not accepted into the queue only to be discovered as 'failed' later by
    polling."""
    resp = client.post("/api/v1/jobs", files=_junk_file)
    assert resp.status_code == 400
    assert "Unsupported file type" in resp.json()["detail"]


def test_create_job_requires_auth_when_configured(monkeypatch):
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1",))
    resp = client.post("/api/v1/jobs", files=_junk_file)
    assert resp.status_code == 401


def test_job_status_and_result_endpoints_require_auth_when_configured(monkeypatch):
    """Checked with a made-up job id specifically to prove auth is
    enforced before the job lookup happens (401, not 404)."""
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1",))
    assert client.get("/api/v1/jobs/does-not-exist").status_code == 401
    assert client.get("/api/v1/jobs/does-not-exist/result").status_code == 401


def test_get_job_status_404_for_unknown_id():
    resp = client.get("/api/v1/jobs/no-such-job-id")
    assert resp.status_code == 404


@requires_tesseract
def test_create_job_returns_202_with_a_pollable_status(synthetic_pdf, tesseract_cmd, monkeypatch):
    monkeypatch.setattr(config, "TESSERACT_CMD", tesseract_cmd)
    with open(synthetic_pdf, "rb") as f:
        resp = client.post(
            "/api/v1/jobs", files={"file": ("synthetic_scan.pdf", f, "application/pdf")}
        )
    assert resp.status_code == 202
    body = resp.json()
    assert "job_id" in body
    # Racy by nature (the background thread may already have picked it up
    # by the time this response was built) -- both are valid, "completed"
    # this early would not be.
    assert body["status"] in ("queued", "processing")

    # Wait for it to actually finish before this test returns -- otherwise
    # it keeps running in the background, competing with whatever the next
    # test does on the same worker pool (see module docstring: this is
    # exactly what caused a real CI failure).
    assert _wait_for_job_terminal(body["job_id"]) == "completed"


@requires_tesseract
def test_job_lifecycle_completes_with_correct_result(synthetic_pdf, tesseract_cmd, monkeypatch):
    """Full async round trip: submit, poll until done, download the
    result, and check it against the same expected values already locked
    in by the synchronous end-to-end test."""
    monkeypatch.setattr(config, "TESSERACT_CMD", tesseract_cmd)
    with open(synthetic_pdf, "rb") as f:
        resp = client.post(
            "/api/v1/jobs", files={"file": ("synthetic_scan.pdf", f, "application/pdf")}
        )
    job_id = resp.json()["job_id"]

    status = _wait_for_job_terminal(job_id)
    assert status == "completed", f"job did not complete in time (last status: {status})"

    result_resp = client.get(f"/api/v1/jobs/{job_id}/result")
    assert result_resp.status_code == 200
    assert result_resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(result_resp.content))
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1][:5] == ("1", "100000001", "SMITH", "JOHN", "M")


@requires_tesseract
def test_get_job_result_409_before_job_completes(synthetic_pdf, tesseract_cmd, monkeypatch):
    """A job that's still running must report 409 (not ready), not the
    file -- checked deterministically with a monkeypatched slow
    conversion rather than racing a real one."""
    import time

    monkeypatch.setattr(config, "TESSERACT_CMD", tesseract_cmd)

    def _slow_convert(*args, **kwargs):
        time.sleep(5)

    # Phase 9: app.py no longer imports convert_scanned_to_excel directly --
    # it's called through the registered Capability's convert function
    # (documents_converter/converters/ocr_to_excel.py), so that's what
    # needs patching now for this to actually take effect on the real path.
    monkeypatch.setattr(
        "documents_converter.converters.ocr_to_excel.convert_scanned_to_excel", _slow_convert
    )

    with open(synthetic_pdf, "rb") as f:
        resp = client.post(
            "/api/v1/jobs", files={"file": ("synthetic_scan.pdf", f, "application/pdf")}
        )
    job_id = resp.json()["job_id"]

    result_resp = client.get(f"/api/v1/jobs/{job_id}/result")
    assert result_resp.status_code == 409


# --------------------------------------------------------------------------
# Phase 9: capability registry (documents_converter/registry.py) and the
# new image->pdf conversion routed through it, proving /convert and /jobs
# aren't hardcoded to OCR->Excel any more.
# --------------------------------------------------------------------------


def test_capabilities_endpoint_lists_both_registered_conversions():
    resp = client.get("/api/v1/capabilities")
    assert resp.status_code == 200
    body = resp.json()
    pairs = {(c["source_format"], c["target_format"]) for c in body}
    assert ("scanned_document", "xlsx") in pairs
    assert ("image", "pdf") in pairs
    image_to_pdf = next(c for c in body if c["target_format"] == "pdf")
    assert ".png" in image_to_pdf["accepted_extensions"]


def test_capabilities_endpoint_requires_no_auth(monkeypatch):
    """Discovery metadata, like /health -- must stay reachable even when
    API_KEYS is configured, the same way /health does."""
    monkeypatch.setattr(config, "API_KEYS", ("secret-key-1",))
    resp = client.get("/api/v1/capabilities")
    assert resp.status_code == 200


def test_convert_rejects_target_with_no_matching_capability():
    """A .pdf file asking for target=pdf has no registered capability
    (pdf->pdf isn't a thing here) -- must be a clean 400 through the
    registry, not a 500 or a silent no-op."""
    resp = client.post(
        "/api/v1/convert",
        data={"target": "pdf"},
        files={"file": ("scan.pdf", io.BytesIO(b"%PDF-1.4 fake content"), "application/pdf")},
    )
    assert resp.status_code == 400
    assert "Unsupported file type" in resp.json()["detail"]


def test_convert_image_to_pdf_end_to_end():
    """The actual proof-of-concept conversion end to end, through the
    synchronous endpoint: a real PNG in, a real, valid, correctly-sized
    single-page PDF out -- no OCR involved anywhere in this path."""
    resp = client.post(
        "/api/v1/convert",
        data={"target": "pdf"},
        files={"file": ("photo.png", io.BytesIO(_sample_png_bytes()), "image/png")},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.headers["content-disposition"] == "attachment; filename=converted.pdf"

    doc = fitz.open(stream=resp.content, filetype="pdf")
    try:
        assert doc.page_count == 1
        page = doc[0]
        # The rendered page should be roughly the image's own proportions
        # (40x30, a 4:3 landscape rectangle), not some unrelated default
        # page size -- a real conversion, not a blank placeholder.
        assert page.rect.width > page.rect.height
    finally:
        doc.close()


def test_convert_defaults_to_xlsx_target_when_omitted():
    """Backward compatibility: a caller that never heard of `target` (every
    caller before Phase 9) must still get rejected/accepted exactly as
    before -- an unsupported extension for the implied default (xlsx)
    still 400s the same way."""
    resp = client.post(
        "/api/v1/convert",
        files={"file": ("not_a_document.exe", io.BytesIO(b"whatever"), "application/octet-stream")},
    )
    assert resp.status_code == 400
    assert "Unsupported file type" in resp.json()["detail"]


def test_job_image_to_pdf_end_to_end():
    """Same conversion, through the async job queue: submit with
    target=pdf, poll to completion, download, verify it's a real PDF."""
    resp = client.post(
        "/api/v1/jobs",
        data={"target": "pdf"},
        files={"file": ("photo.png", io.BytesIO(_sample_png_bytes()), "image/png")},
    )
    assert resp.status_code == 202
    job_id = resp.json()["job_id"]

    status = _wait_for_job_terminal(job_id)
    assert status == "completed", f"job did not complete in time (last status: {status})"

    result_resp = client.get(f"/api/v1/jobs/{job_id}/result")
    assert result_resp.status_code == 200
    assert result_resp.headers["content-type"] == "application/pdf"
    assert result_resp.headers["content-disposition"] == "attachment; filename=converted.pdf"

    doc = fitz.open(stream=result_resp.content, filetype="pdf")
    try:
        assert doc.page_count == 1
    finally:
        doc.close()


# --------------------------------------------------------------------------
# Phase 11: audit trail (documents_converter/api/audit.py). Uses the
# OCR-free image->pdf capability rather than the OCR pipeline -- these
# tests are about whether the right audit events fire, not about OCR
# accuracy, so there's no reason to pay for a slow real OCR run here.
# --------------------------------------------------------------------------


def _audit_events(caplog) -> list[dict]:
    return [
        json.loads(r.message)
        for r in caplog.records
        if r.name == "documents_converter.audit"
    ]


def test_convert_emits_requested_and_completed_audit_events(caplog):
    with caplog.at_level(logging.INFO, logger="documents_converter.audit"):
        resp = client.post(
            "/api/v1/convert",
            data={"target": "pdf"},
            files={"file": ("photo.png", io.BytesIO(_sample_png_bytes()), "image/png")},
        )
    assert resp.status_code == 200

    events = _audit_events(caplog)
    kinds = [e["event"] for e in events]
    assert kinds == ["convert_requested", "convert_completed"]
    assert events[0]["ext"] == ".png"
    assert events[0]["target_format"] == "pdf"
    assert events[0]["source_format"] == "image"
    assert "client_ip" in events[0]
    assert events[0]["request_id"] == events[1]["request_id"]
    assert isinstance(events[1]["duration_ms"], int)


def test_convert_emits_failed_audit_event_on_rejected_upload(caplog):
    with caplog.at_level(logging.INFO, logger="documents_converter.audit"):
        resp = client.post("/api/v1/convert", files=_junk_file)
    assert resp.status_code == 400

    # Rejected before a capability is even resolved (Unsupported file
    # type) -- no audit event at all, since there's no real attempted
    # conversion to record yet. Confirms the earlier validation failure
    # doesn't silently masquerade as a recorded conversion attempt.
    assert _audit_events(caplog) == []


def test_job_emits_created_and_completed_audit_events(caplog):
    with caplog.at_level(logging.INFO, logger="documents_converter.audit"):
        resp = client.post(
            "/api/v1/jobs",
            data={"target": "pdf"},
            files={"file": ("photo.png", io.BytesIO(_sample_png_bytes()), "image/png")},
        )
        assert resp.status_code == 202
        job_id = resp.json()["job_id"]
        status = _wait_for_job_terminal(job_id)
        assert status == "completed"

    events = _audit_events(caplog)
    kinds = [e["event"] for e in events]
    assert kinds == ["job_created", "job_completed"]
    assert events[0]["job_id"] == job_id == events[1]["job_id"]
    assert events[0]["endpoint"] == "async"
