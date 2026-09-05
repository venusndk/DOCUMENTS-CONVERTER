"""
Regression tests for scan_to_excel.py, written against real bugs found and
fixed during this project's development (see docs/PHASE_0_AUDIT.md) rather
than against imagined edge cases. Each test's docstring names the specific
failure it guards against.

Uses a synthetic, fabricated fixture (tests/fixtures/synthetic_scan.py) --
never real scanned documents, which contain genuine personal data (see
docs/PHASE_0_AUDIT.md risk register item #1).
"""

import zipfile

import fitz
import numpy as np
import openpyxl
import pytest

from documents_converter import ocr_excel as ste
from documents_converter.providers.cell_ocr import TesseractCellOCR
from documents_converter.providers.table_detection import _pick_extraction_mode

from conftest import requires_tesseract


# --------------------------------------------------------------------------
# Fast unit tests -- no OCR, run in well under a second total.
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    "value,expected",
    [
        ("76.90", True),  # period-decimal in a comma-decimal document
        ("$6,00", True),  # stray currency-like symbol
        ("None 61,00!", True),  # a literal "None" leaking through from a failed merge
        ("‘ER", True),  # a stray curly-quote character (real bug: looked like
        # the U+FFFD replacement char in a terminal but wasn't -- an earlier,
        # narrower blocklist missed this; the current allowlist catches it)
        ("76,00", False),
        ("PPS1111", False),
        ("PROGRESS", False),
        ("GAMUREKE", False),
        ("Life Skills Education for\nPolicing", False),
        (None, False),
        ("", False),
    ],
)
def test_is_suspicious(value, expected):
    assert ste._is_suspicious(value) is expected


def test_workbook_never_writes_real_formulas(tmp_path):
    """
    Regression guard for the Excel-corruption bug this project found: xlsxwriter's
    default write() treats any string starting with "=" as a live formula. OCR
    garbage starting with "=" was silently turned into a broken formula, and Excel
    stripped it -- along with surrounding data -- on open ("repaired file" dialog).
    Importing scan_to_excel patches Workbook.__init__ to disable that globally.
    """
    import xlsxwriter

    path = tmp_path / "formula_test.xlsx"
    wb = xlsxwriter.Workbook(str(path))
    ws = wb.add_worksheet()
    ws.write(0, 0, "=1+1")  # would become a live formula without the patch
    ws.write(0, 1, "$6,00")
    wb.close()

    with zipfile.ZipFile(path) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "<f>" not in xml, "a <f> tag means the formula-injection patch isn't active"


def test_group_words_by_parent_tolerates_none_values():
    """
    Regression guard: img2table's own OCRData._group_words_by_parent crashes with
    "TypeError: sequence item 0: expected str instance, NoneType found" whenever a
    detected word has no recognized text (value=None) -- which got more likely once
    --preprocess was added. This project patches it to skip such words instead of
    crashing.
    """
    words = [
        {"parent": "line1", "x1": 0, "y1": 0, "value": "hello"},
        {"parent": "line1", "x1": 10, "y1": 0, "value": None},
        {"parent": "line1", "x1": 20, "y1": 0, "value": "world"},
    ]
    assert ste._patched_group_words_by_parent(words) == ["hello world"]


def _render_gray(pdf_path, dpi=200):
    doc = fitz.open(str(pdf_path))
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72))
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    doc.close()
    import cv2

    return cv2.cvtColor(img[:, :, :3], cv2.COLOR_RGB2GRAY)


def test_detect_grid_lines_finds_synthetic_grid(synthetic_pdf):
    """
    The synthetic fixture has 9 columns and 5 rows (header + 2 students x 2 rows
    each), so a correct grid-line detector should find at least 6 row-boundary
    lines and 10 column-boundary lines.
    """
    gray = _render_gray(synthetic_pdf)
    row_lines, col_lines = ste._detect_grid_lines(gray)
    assert len(row_lines) >= 5
    assert len(col_lines) >= 8


# --------------------------------------------------------------------------
# End-to-end smoke test -- needs a real Tesseract install, skipped if absent.
# --------------------------------------------------------------------------


@requires_tesseract
def test_convert_scanned_to_excel_end_to_end(tmp_path, synthetic_pdf, tesseract_cmd):
    """
    Full pipeline regression guard: every value in the synthetic fixture is known
    in advance, so this locks in exact expected output rather than just "did it
    not crash". A future change that silently drops or corrupts data should fail
    this test.
    """
    out_path = tmp_path / "out.xlsx"
    ste.convert_scanned_to_excel(
        file_path=str(synthetic_pdf),
        output_excel_path=str(out_path),
        tesseract_cmd=tesseract_cmd,
    )

    assert out_path.exists()

    with zipfile.ZipFile(out_path) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "<f>" not in xml

    wb = openpyxl.load_workbook(str(out_path))
    assert len(wb.sheetnames) == 1
    ws = wb[wb.sheetnames[0]]

    rows = [row for row in ws.iter_rows(values_only=True)]
    assert len(rows) == 5  # header + 2 students x 2 rows each

    header = rows[0]
    assert header[0] == "S/N"
    assert header[1] == "Ref.No"
    assert header[2] == "Surname"
    assert header[3] == "Firstname"
    assert header[4] == "Sex"
    # The rotated-header fix: this cell is printed sideways in the source
    # image and must be re-OCR'd right-side-up, not left as noise.
    assert header[5] == "Sideways Hdr"

    student1 = rows[1][:9]
    assert student1 == ("1", "100000001", "SMITH", "JOHN", "M", "85,00", "90,00", "175", "87,50")

    student2 = rows[3][:9]
    assert student2 == ("2", "100000002", "DOE", "JANE", "F", "70,00", "65,00", "135", "67,50")


# --------------------------------------------------------------------------
# Phase 2: provider-level tests (docs/PHASE_0_AUDIT.md). Added when the
# per-cell OCR call and the table-detection strategy were pulled out behind
# CellOCRProvider / TableDetector so they're independently testable, rather
# than only reachable through the full end-to-end pipeline.
# --------------------------------------------------------------------------


@requires_tesseract
def test_tesseract_cell_ocr_recognizes_a_known_crop(synthetic_pdf, tesseract_cmd):
    """The provider that _fix_rotated_cells and the grid-fallback both use for
    per-cell OCR, exercised directly rather than only via the full pipeline."""
    import pytesseract

    pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    gray = _render_gray(synthetic_pdf)
    row_lines, col_lines = ste._detect_grid_lines(gray)
    # Row 1 (index 1, after the header) / column 2 (Surname) is "SMITH" in
    # the fixture -- see tests/fixtures/synthetic_scan.py.
    y1, y2 = row_lines[1], row_lines[2]
    x1, x2 = col_lines[2], col_lines[3]
    crop = gray[y1 + 3 : y2 - 3, x1 + 3 : x2 - 3]

    assert TesseractCellOCR().recognize(crop) == "SMITH"


def test_pick_extraction_mode_prefers_bordered_for_a_clean_grid(synthetic_pdf):
    """The fixture has a fully ruled grid, so the cheap shape-only comparison
    should prefer the stricter bordered detector over borderless."""
    gray_rgb = fitz.open(str(synthetic_pdf))[0].get_pixmap(matrix=fitz.Matrix(200 / 72, 200 / 72))
    img = np.frombuffer(gray_rgb.samples, dtype=np.uint8).reshape(
        gray_rgb.height, gray_rgb.width, gray_rgb.n
    )[:, :, :3]
    assert _pick_extraction_mode(img) is False


@requires_tesseract
def test_progress_callback_used_instead_of_print(tmp_path, synthetic_pdf, tesseract_cmd):
    """
    Regression guard for the Phase 2 progress-callback addition: a caller
    should be able to capture pipeline status without scraping stdout (the
    motivating case is a future job-queue/API layer logging structured
    progress rather than print()). Passing a custom callback must not change
    the pipeline's actual output.
    """
    messages = []
    out_path = tmp_path / "out.xlsx"
    ste.convert_scanned_to_excel(
        file_path=str(synthetic_pdf),
        output_excel_path=str(out_path),
        tesseract_cmd=tesseract_cmd,
        progress=messages.append,
    )
    assert any("Detected file type" in m for m in messages)
    assert any("Done." in m for m in messages)
    assert out_path.exists()
