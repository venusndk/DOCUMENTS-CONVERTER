"""
Real browser-level tests for the Phase 8 frontend
(documents_converter/api/static/index.html).

Uses Playwright to actually load the page in a headless browser and
interact with it the way a real person would -- choose a file, click
Convert, wait for the real download -- rather than only asserting the
HTML contains expected markup. The whole point of Phase 8 is that a
non-technical person can use this without curl; a test that doesn't
drive a real browser wouldn't actually prove that.
"""

from __future__ import annotations

import threading
import time

import openpyxl
import pytest
import requests
import uvicorn

from documents_converter.api import config
from documents_converter.api.app import app

from conftest import requires_tesseract

_PORT = 8765


@pytest.fixture(scope="module")
def running_app_server(tesseract_cmd):
    """
    Playwright drives a real browser against a real URL -- it can't talk
    to FastAPI's in-process TestClient the way the other test files do.
    Runs uvicorn in a background thread for the module's test session.
    """
    config.TESSERACT_CMD = tesseract_cmd
    server = uvicorn.Server(
        uvicorn.Config(app, host="127.0.0.1", port=_PORT, log_level="warning")
    )
    thread = threading.Thread(target=server.run, daemon=True)
    thread.start()

    deadline = time.monotonic() + 10
    base_url = f"http://127.0.0.1:{_PORT}"
    while time.monotonic() < deadline:
        try:
            if requests.get(f"{base_url}/health", timeout=1).status_code == 200:
                break
        except requests.exceptions.ConnectionError:
            time.sleep(0.2)
    else:
        raise RuntimeError("live server did not start in time")

    yield base_url
    server.should_exit = True


def test_index_page_loads_with_expected_elements(running_app_server):
    """Fast, non-Playwright sanity check -- the full browser test below is
    slower and OCR-dependent, this one just confirms the route itself is
    wired correctly."""
    resp = requests.get(running_app_server + "/")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/html")
    assert 'id="submit-btn"' in resp.text
    assert 'id="file-input"' in resp.text


@requires_tesseract
def test_frontend_full_upload_to_download_flow(running_app_server, synthetic_pdf, tmp_path):
    """
    The real end-to-end proof: a real headless browser loads the actual
    page, selects the actual synthetic fixture through the actual file
    input, clicks the actual Convert button, and the actual polling/
    download JavaScript in index.html produces a real file -- checked
    against the same expected values as every other end-to-end test in
    this project.
    """
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            page = browser.new_page()
            page.goto(running_app_server + "/")
            assert page.title() == "Documents Converter"

            page.set_input_files("#file-input", str(synthetic_pdf))
            submit = page.locator("#submit-btn")
            assert not submit.is_disabled()

            with page.expect_download(timeout=60_000) as download_info:
                submit.click()
            download = download_info.value

            saved_path = tmp_path / "downloaded.xlsx"
            download.save_as(str(saved_path))
        finally:
            browser.close()

    wb = openpyxl.load_workbook(saved_path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1][:5] == ("1", "100000001", "SMITH", "JOHN", "M")


@requires_tesseract
def test_frontend_shows_error_for_rejected_upload(running_app_server, tmp_path):
    """A file the API rejects (bad extension) should surface a visible
    error in the page, not fail silently or hang."""
    from playwright.sync_api import sync_playwright

    bad_file = tmp_path / "not_a_document.exe"
    bad_file.write_bytes(b"whatever")

    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            page = browser.new_page()
            page.goto(running_app_server + "/")
            page.set_input_files("#file-input", str(bad_file))
            page.locator("#submit-btn").click()
            status = page.locator("#status")
            page.wait_for_function(
                "document.getElementById('status').className === 'error'", timeout=10_000
            )
            assert "Unsupported file type" in status.inner_text()
        finally:
            browser.close()
